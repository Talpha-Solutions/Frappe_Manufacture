# Copyright (c) 2026, talpha solutions and contributors
# For license information, please see license.txt

from __future__ import annotations

import io
from typing import Any

import frappe

from fitzgerald_kitchens.workbook_import.parser import WorkbookParseError


class WorkbookReader:
	def __init__(self, file_url: str):
		self.file_url = file_url
		self._workbook = None
		self._sheet_names: list[str] = []

	def load(self) -> None:
		content, extension = self._read_file_content()
		if extension not in ("xlsx", "xls"):
			raise WorkbookParseError(
				frappe._("Full workbook import requires an XLSX file. Use Quantity Only mode for CSV.")
			)
		try:
			import openpyxl
		except ImportError as exc:
			raise WorkbookParseError(frappe._("XLSX import requires openpyxl.")) from exc

		self._workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
		self._sheet_names = list(self._workbook.sheetnames)

	def sheet_names(self) -> list[str]:
		return list(self._sheet_names)

	def has_sheet(self, name: str) -> bool:
		return name in self._sheet_names

	def get_sheet_rows(self, sheet_name: str) -> list[list[Any]]:
		if not self._workbook or sheet_name not in self._workbook.sheetnames:
			raise WorkbookParseError(frappe._("Sheet '{0}' not found in workbook.").format(sheet_name))

		ws = self._workbook[sheet_name]
		return [list(row) for row in ws.iter_rows(values_only=True)]

	def get_sheet_as_dicts(self, sheet_name: str, header_row_index: int = 0) -> list[dict[str, Any]]:
		rows = self.get_sheet_rows(sheet_name)
		if not rows or header_row_index >= len(rows):
			return []

		headers = [_stringify_cell(cell) for cell in rows[header_row_index]]
		result: list[dict[str, Any]] = []

		for row_idx, values in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
			if not any(_stringify_cell(v) for v in values):
				continue
			row_dict: dict[str, Any] = {"row_number": row_idx}
			for col_idx, header in enumerate(headers):
				if not header:
					continue
				value = values[col_idx] if col_idx < len(values) else None
				row_dict[header] = _stringify_cell(value)
			result.append(row_dict)

		return result

	def _read_file_content(self) -> tuple[bytes, str]:
		file_doc = frappe.get_doc("File", {"file_url": self.file_url})
		content = file_doc.get_content()
		extension = (file_doc.file_name or self.file_url).rsplit(".", 1)[-1].lower()
		return content, extension


def _stringify_cell(value: Any) -> str:
	if value is None:
		return ""
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	return str(value).strip()
